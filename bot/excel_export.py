"""Вивантаження замовлень / історії балансу в Excel (.xlsx)."""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def order_history_bucket(order: dict[str, Any]) -> str:
    """Як у miniapp: awaiting | transit | received | returns."""
    payload = order.get("payload") or {}
    ret = payload.get("dropper_return")
    ttn = str(order.get("ttn_status") or "")
    if isinstance(ret, dict) and ret:
        return "returns"
    if ttn in {
        "returned",
        "refused",
        "return_at_warehouse",
        "cancelled",
    } or payload.get("return_at_warehouse") or str(order.get("status") or "") == "cancelled":
        return "returns"
    if ttn == "received":
        return "received"
    if ttn in {"in_transit", "at_warehouse", "provided"}:
        return "transit"
    return "awaiting"


def bucket_label_uk(bucket: str) -> str:
    return {
        "awaiting": "Очікує відправлення",
        "transit": "В дорозі",
        "received": "Отримано",
        "returns": "Повернення",
    }.get(bucket, bucket or "—")


def _order_created_day(order: dict[str, Any]) -> str:
    raw = str(order.get("created_at") or "").strip()
    if not raw:
        return ""
    try:
        if raw.endswith("Z"):
            raw = raw[:-1] + "+00:00"
        dt = datetime.fromisoformat(raw)
        return dt.date().isoformat()
    except ValueError:
        return raw[:10] if len(raw) >= 10 else ""


def order_matches_export_filters(
    order: dict[str, Any],
    *,
    status: str = "",
    date_from: str = "",
    date_to: str = "",
) -> bool:
    st = str(status or "").strip().lower()
    if st and st not in {"", "all"}:
        if order_history_bucket(order) != st:
            return False
    day = _order_created_day(order)
    if date_from and (not day or day < date_from):
        return False
    if date_to and (not day or day > date_to):
        return False
    return True


def _cart_summary(cart: Any) -> str:
    if not isinstance(cart, list):
        return ""
    parts: list[str] = []
    for item in cart:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code") or "").strip()
        name = str(item.get("name") or "").strip()
        color = str(item.get("color") or "").strip()
        try:
            qty = int(item.get("qty") or item.get("quantity") or 1)
        except (TypeError, ValueError):
            qty = 1
        bit = code or name or "товар"
        if color:
            bit = f"{bit} ({color})"
        if qty != 1:
            bit = f"{bit} ×{qty}"
        parts.append(bit)
    return "; ".join(parts)


def _recipient_name(payload: dict[str, Any]) -> str:
    r = payload.get("recipient") or {}
    return " ".join(
        p
        for p in (
            str(r.get("last_name") or "").strip(),
            str(r.get("first_name") or "").strip(),
            str(r.get("patronymic") or "").strip(),
        )
        if p
    )


def order_export_row(order: dict[str, Any]) -> dict[str, Any]:
    payload = order.get("payload") or {}
    recipient = payload.get("recipient") or {}
    delivery = payload.get("delivery") or {}
    payment = payload.get("payment") or {}
    bucket = order_history_bucket(order)
    return {
        "order_number": order.get("order_number") or "",
        "created_at": order.get("created_at") or "",
        "status_bucket": bucket_label_uk(bucket),
        "ttn_status": order.get("ttn_status") or "",
        "ttn_number": order.get("ttn_number") or payload.get("ttn_number") or "",
        "client_name": _recipient_name(payload),
        "client_phone": recipient.get("phone") or "",
        "city": delivery.get("city") or "",
        "warehouse": delivery.get("warehouse") or "",
        "delivery_method": order.get("delivery_method") or delivery.get("method") or "",
        "payment_method": order.get("payment_method") or payment.get("method") or "",
        "drop_total": order.get("total"),
        "prepay": order.get("prepay"),
        "cod_amount": order.get("cod_amount"),
        "cart": _cart_summary(payload.get("cart")),
        "comment": payload.get("comment") or "",
        "own_ttn": "так" if order.get("own_ttn") or payload.get("own_ttn") else "ні",
    }


ORDER_HEADERS = [
    ("order_number", "№ замовлення"),
    ("created_at", "Дата"),
    ("status_bucket", "Статус"),
    ("ttn_status", "Статус ТТН"),
    ("ttn_number", "ТТН"),
    ("client_name", "Клієнт"),
    ("client_phone", "Телефон"),
    ("city", "Місто"),
    ("warehouse", "Відділення"),
    ("delivery_method", "Доставка"),
    ("payment_method", "Оплата"),
    ("drop_total", "Дроп ціна"),
    ("prepay", "Передплата"),
    ("cod_amount", "Накладний"),
    ("cart", "Товари"),
    ("comment", "Коментар"),
    ("own_ttn", "Власна ТТН"),
]

LEDGER_HEADERS = [
    ("created_at", "Дата операції"),
    ("entry_type", "Тип операції"),
    ("title", "Назва"),
    ("amount", "Сума"),
    ("note", "Примітка"),
    ("related_order_id", "№ замовлення (ledger)"),
]


def build_orders_xlsx(
    orders: list[dict[str, Any]],
    *,
    sheet_title: str = "Замовлення",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Замовлення"
    header_font = Font(bold=True)
    for col, (_key, label) in enumerate(ORDER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
    for r_idx, order in enumerate(orders, start=2):
        row = order_export_row(order)
        for c_idx, (key, _label) in enumerate(ORDER_HEADERS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(key))
    _autosize(ws, len(ORDER_HEADERS))
    return _wb_bytes(wb)


def build_balance_ledger_xlsx(
    rows: list[dict[str, Any]],
    *,
    sheet_title: str = "Баланс",
) -> bytes:
    """
    rows: dict з ключами ledger + вкладений order (або плоскі order_* поля).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Баланс"
    headers = LEDGER_HEADERS + ORDER_HEADERS
    header_font = Font(bold=True)
    for col, (_key, label) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
    for r_idx, item in enumerate(rows, start=2):
        ledger = item.get("ledger") or item
        order = item.get("order")
        order_row = order_export_row(order) if order else {k: "" for k, _ in ORDER_HEADERS}
        flat = {
            "created_at": ledger.get("created_at") or "",
            "entry_type": ledger.get("entry_type") or "",
            "title": ledger.get("title") or "",
            "amount": ledger.get("amount"),
            "note": ledger.get("note") or "",
            "related_order_id": ledger.get("related_order_id") or "",
            **order_row,
        }
        for c_idx, (key, _label) in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=flat.get(key))
    _autosize(ws, len(headers))
    return _wb_bytes(wb)


def _autosize(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = 10
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=min(ws.max_row, 200)):
            for cell in row:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = min(len(val), 48)
        ws.column_dimensions[letter].width = max_len + 2


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def safe_filename(prefix: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    clean = re.sub(r"[^\w\-]+", "_", prefix, flags=re.UNICODE).strip("_") or "export"
    return f"{clean}_{stamp}.xlsx"
